"""
Réplica fiel (Python) de la lectura de RUT del Validador Acreditable, para
auditar los documentos reales de Antofagasta fuera del navegador — mismo
método que ya usa el resto del skill acreditable-review.

Portado línea por línea desde index.html:
  - va_normRut / va_calcDV / va_addRut / va_findAllRuts (~10650-10725)
  - va_dvRut / va_rutValido / va_extraerRutConChecksum (~10744-10850)
  - va_getPdfTextOCR (~12697): texto nativo por página, OCR (Tesseract,
    PSM 3 default) solo si texto nativo <40 chars y total<=maxPagesOCR.

Requiere: pip install pymupdf pytesseract openpyxl pillow
Tesseract binario: C:\\Users\\agutierrez\\AppData\\Local\\Programs\\Tesseract-OCR\\tesseract.exe
"""
import re, sys, io
import fitz
import pytesseract
from PIL import Image
import openpyxl

pytesseract.pytesseract.tesseract_cmd = r"C:\Users\agutierrez\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"

# ── va_calcDV / va_addRut / va_findAllRuts ──
def calc_dv(body):
    d = re.sub(r'\D', '', body)
    s = 0
    mul = [2,3,4,5,6,7]
    j = 0
    for i in range(len(d)-1, -1, -1):
        s += int(d[i]) * mul[j % 6]
        j += 1
    r = 11 - (s % 11)
    if r == 11: return '0'
    if r == 10: return 'k'
    return str(r)

def norm_rut(r):
    s = re.sub(r'[.\s,|]', '', str(r or '')).lower().strip()
    if len(s) >= 2 and '-' not in s:
        s = s[:-1] + '-' + s[-1]
    return s

def add_rut(rutset, r):
    n = norm_rut(r)
    if len(n) < 9: return
    parts = n.split('-')
    if len(parts) == 2:
        body = re.sub(r'\D', '', parts[0])
        dv = parts[1].lower()
        if calc_dv(body) == dv:
            rutset.add(n)

def find_all_ruts(txt):
    ruts = set()
    # Patrón 1: con candado anti-monto (?<![\d.,])
    for m in re.finditer(r'(?<![\d.,])(\d{1,2})[.,\s](\d{3})[.,\s](\d{3})[\s-]+(\d|[kK])', txt):
        add_rut(ruts, m.group(1)+m.group(2)+m.group(3)+'-'+m.group(4))
    # Patrón 2: sin separadores
    for m in re.finditer(r'(?<!\d)(\d{7,8})\s*-\s*(\d|[kK])(?!\d)', txt):
        add_rut(ruts, m.group(1)+'-'+m.group(2))
    # Patrón 3: "Rut, 16,383,943-1"
    for m in re.finditer(r'Rut[,\s]+(\d{1,2})[,](\d{3})[,](\d{3})\s*[-]\s*([\dkK])', txt, re.I):
        add_rut(ruts, m.group(1)+m.group(2)+m.group(3)+'-'+m.group(4))
    # Patrón 4: "RUT:" + dígitos
    for m in re.finditer(r'RUT[:\s.]*(\d[\d\s|.,/-]{6,20}[\dkK])', txt, re.I):
        cleaned = re.sub(r'[\s|.,]', '', m.group(1))
        add_rut(ruts, cleaned)
    # Patrón 5: "Número de Documento"/"Cedula"/"identidad"
    for m in re.finditer(r'(?:N[uú]mero de Documento|Cedula[^:]*N[ºo]?|identidad\s*N[ºo]?)[:\s]*(\d[\d\s|.,/-]{6,16}[\dkK])', txt, re.I):
        cleaned = re.sub(r'[\s|.,]', '', m.group(1))
        add_rut(ruts, cleaned)
    # Patrón 6: 7 dígitos sueltos
    for m in re.finditer(r'\b(\d)\s(\d)\s(\d)\s(\d)\s(\d)\s(\d)\s(\d)\s*[-–]\s*(\d|[kK])\b', txt):
        add_rut(ruts, ''.join(m.groups()[:7])+'-'+m.group(8))
    # Patrón 7: 8 dígitos sueltos
    for m in re.finditer(r'\b(\d)\s(\d)\s(\d)\s(\d)\s(\d)\s(\d)\s(\d)\s(\d)\s*[-–]\s*(\d|[kK])\b', txt):
        add_rut(ruts, ''.join(m.groups()[:8])+'-'+m.group(9))
    # Patrón 8: boxes con |
    for m in re.finditer(r'[|]?\s*(\d)\s*[|]\s*(\d)\s*[|]\s*(\d)\s*[|]\s*(\d)\s*[|]\s*(\d)\s*[|]\s*(\d)\s*[|]\s*(\d)\s*[|]\s*[-]\s*[|]\s*(\d|[kK])\s*[|]?', txt):
        add_rut(ruts, ''.join(m.groups()[:7])+'-'+m.group(8))
    return ruts

# ── va_findAllRutsRaw: mismo patrón 1/2/4 SIN validar DV, para alimentar el
# rescate por Levenshtein (va_matchRutCercano) cuando el checksum descarta
# el candidato real por 1 dígito mal leído ──
def find_all_ruts_raw(txt):
    ruts = set()
    for m in re.finditer(r'(?<![\d.,])(\d{1,2})[.,\s](\d{3})[.,\s](\d{3})[\s-]+(\d|[kK])', txt):
        ruts.add(norm_rut(m.group(1)+m.group(2)+m.group(3)+'-'+m.group(4)))
    for m in re.finditer(r'(?<!\d)(\d{7,8})\s*-\s*(\d|[kK])(?!\d)', txt):
        ruts.add(norm_rut(m.group(1)+'-'+m.group(2)))
    for m in re.finditer(r'RUT[:\s.]*(\d[\d\s|.,/-]{6,20}[\dkK])', txt, re.I):
        ruts.add(norm_rut(re.sub(r'[\s|.,]', '', m.group(1))))
    return ruts

def match_rut_cercano(candidato, lista_validos):
    mejor=None; mejor_d=999; segundo_d=999
    for r in lista_validos:
        d = levenshtein(candidato, r)
        if d < mejor_d: segundo_d = mejor_d; mejor_d = d; mejor = r
        elif d < segundo_d: segundo_d = d
    if mejor is None or mejor_d > 1: return None
    if segundo_d - mejor_d < 1: return None
    return mejor

# ── va_dvRut / va_extraerRutConChecksum (rescate por checksum) ──
def dv_rut(cuerpo):
    suma = 0; mult = 2
    for ch in reversed(cuerpo):
        suma += int(ch) * mult
        mult = 2 if mult == 7 else mult + 1
    resto = 11 - (suma % 11)
    if resto == 11: return '0'
    if resto == 10: return 'K'
    return str(resto)

def rut_valido(cuerpo, dv):
    if not re.fullmatch(r'\d{7,8}', cuerpo): return False
    return dv_rut(cuerpo) == str(dv).upper()

def extraer_rut_con_checksum(linea, lista_validos_norm):
    solo_dig = re.sub(r'[^0-9kK]', '', linea or '')
    candidatos = []
    for length in (8, 9):
        for i in range(0, max(0, len(solo_dig)-length+1)):
            chunk = solo_dig[i:i+length]
            cuerpo, dv = chunk[:-1], chunk[-1]
            if rut_valido(cuerpo, dv):
                n = norm_rut(cuerpo+'-'+dv)
                if n in lista_validos_norm:
                    candidatos.append(n)
    return candidatos[0] if len(set(candidatos)) == 1 else None

# ── va_getPdfTextOCR: texto nativo, OCR (PSM 3 default) si <40 chars y total<=maxPagesOCR ──
def get_pdf_text_ocr(path, max_pages_ocr=30, intentar_rotacion=False, psm=None, scale=2.5, verbose=True):
    doc = fitz.open(path)
    total = doc.page_count
    all_text = []
    diag_nativo = diag_intentado = diag_exitoso = 0
    for i in range(total):
        page = doc[i]
        txt = page.get_text()
        txt_limpio = re.sub(r'\s', '', txt)
        if len(txt_limpio) < 40 and total <= max_pages_ocr:
            diag_intentado += 1
            try:
                mat = fitz.Matrix(scale, scale)
                pix = page.get_pixmap(matrix=mat)
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                config = f'--psm {psm}' if psm else ''
                ocr_txt = pytesseract.image_to_string(img, lang='spa', config=config)
                if ocr_txt and len(re.sub(r'\s','',ocr_txt)) > 0:
                    txt = ocr_txt
                    diag_exitoso += 1
            except Exception as e:
                print(f'  [OCR failed pág {i+1}] {e}')
        else:
            diag_nativo += 1
        all_text.append(txt)
    full = '\n'.join(all_text)
    if verbose:
        print(f'  [get_pdf_text_ocr] {path.split(chr(92))[-1]}: páginas={total} maxPagesOCR={max_pages_ocr} (total<=tope? {total<=max_pages_ocr}) | nativo-suficiente={diag_nativo} OCR-intentado={diag_intentado} OCR-exitoso={diag_exitoso} | texto total={len(re.sub(chr(92)+"s","",full))} chars')
    return full, total

# ── Nómina (LH) ──
def cargar_lh(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column+1):
        h = ws.cell(row=6, column=c).value
        if h: headers[h.strip()] = c
    def col(name_frag):
        for h,c in headers.items():
            if name_frag.lower() in h.lower():
                return c
        return None
    c_rut = col('mero de Documento') or col('Numero de Documento')
    c_nombre = col('Nombre Completo')
    c_ingreso = col('Fecha Ingreso Compa')
    c_termino = col('Fecha T')
    c_sexo = col('Sexo')
    c_disc = col('Situaci') and [c for h,c in headers.items() if 'Discapacidad' in h]
    c_disc = c_disc[0] if c_disc else None
    c_jub = col('Jubilado')
    rows = []
    for r in range(7, ws.max_row+1):
        rut = ws.cell(row=r, column=c_rut).value if c_rut else None
        if not rut: continue
        nombre = ws.cell(row=r, column=c_nombre).value if c_nombre else ''
        termino = ws.cell(row=r, column=c_termino).value if c_termino else None
        sexo = ws.cell(row=r, column=c_sexo).value if c_sexo else None
        disc = ws.cell(row=r, column=c_disc).value if c_disc else None
        jub = ws.cell(row=r, column=c_jub).value if c_jub else None
        rows.append({
            'rut': str(rut), 'rutNorm': norm_rut(str(rut)), 'nombre': nombre,
            'termino': bool(termino), 'sexo': sexo, 'discapacidad': bool(disc) and str(disc).lower() not in ('no','false','0',''),
            'jubilado': bool(jub) and str(jub).lower() not in ('no','false','0','')
        })
    return rows

# ── va_normalizarNombre / va_tokensNombre / va_coberturaTokens / va_matchNombreNominaConMotivo ──
import unicodedata

VA_STOP_TOKENS = {'SR','SRA','DON','DONA','SENOR','SENORA','DE','LA','EL',
                   'MES','DIA','DIAS','ANO','ANIO','TOTAL','HORAS','NOTA'}

def normalizar_nombre(s):
    s = (s or '').upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def tokens_nombre(s):
    norm = normalizar_nombre(s)
    out = []
    for t in norm.split(' '):
        if t and len(t) >= 2 and t not in VA_STOP_TOKENS and t not in out:
            out.append(t)
    return out

def levenshtein(a, b):
    m, n = len(a), len(b)
    dp = [[0]*(n+1) for _ in range(m+1)]
    for i in range(m+1): dp[i][0] = i
    for j in range(n+1): dp[0][j] = j
    for i in range(1, m+1):
        for j in range(1, n+1):
            dp[i][j] = dp[i-1][j-1] if a[i-1]==b[j-1] else 1+min(dp[i-1][j-1],dp[i-1][j],dp[i][j-1])
    return dp[m][n]

def token_match(t1, t2):
    if t1 == t2: return True
    d = levenshtein(t1, t2)
    tol = 1 if min(len(t1), len(t2)) <= 5 else 2
    return d <= tol

def cobertura_tokens(a_tokens, b_tokens):
    if not a_tokens: return 0
    matched = 0
    usados = set()
    for lt in a_tokens:
        for i, bt in enumerate(b_tokens):
            if i in usados: continue
            if token_match(lt, bt):
                matched += 1
                usados.add(i)
                break
    return matched / len(a_tokens)

def match_nombre_nomina_con_motivo(nombre_leido, lista_lh):
    lt = tokens_nombre(nombre_leido)
    if not lt or not lista_lh:
        return None, 'sin_tokens', []
    scored = []
    for t in lista_lh:
        ht = tokens_nombre(t['nombre'])
        if not ht: continue
        scored.append((cobertura_tokens(lt, ht), t))
    if not scored:
        return None, 'sin_candidatos', []
    scored.sort(key=lambda x: -x[0])
    mejor_cov, mejor_t = scored[0]
    segundo_cov = scored[1][0] if len(scored) > 1 else 0
    if len(lt) >= 2:
        if mejor_cov < 1: return None, 'sin_cobertura', []
    else:
        if mejor_cov < 1 or len(lt[0]) < 4: return None, 'sin_cobertura', []
    if mejor_cov - segundo_cov < 0.34:
        empatados = [t for cov, t in scored if cov >= mejor_cov - 1e-9]
        if len(empatados) > 1:
            return None, 'ambiguo', [e['nombre'] for e in empatados]
    return mejor_t, 'ok', []


if __name__ == '__main__':
    lh_path = sys.argv[1] if len(sys.argv) > 1 else None
    if lh_path:
        rows = cargar_lh(lh_path)
        print(f'LH cargado: {len(rows)} trabajadores')
        print('mujeres:', sum(1 for x in rows if x['sexo']=='F'))
        print('discapacidad:', sum(1 for x in rows if x['discapacidad']))
        print('termino:', sum(1 for x in rows if x['termino']))
