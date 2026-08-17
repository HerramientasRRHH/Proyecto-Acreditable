import sys, re, io
sys.path.insert(0, r'.claude\skills\acreditable-review\scripts')
from audit_antofagasta import *
import openpyxl

VA_LIC_TABLA_ROW_RE = re.compile(
    r'(\d{1,2}\.?\d{3}\.?\d{3}-[\dkK])\s+([A-Za-zÁÉÍÓÚÑñ][A-Za-zÁÉÍÓÚÑñ\s]{4,45}?)\s+(\d{1,2}-\d{1,2}-\d{4})\s+(\d{1,2}-\d{1,2}-\d{4})\s+(\d{1,3})\b'
)

wb = openpyxl.load_workbook(r'Documentos Antofagasta\2026-08-17 Libro de haberes Revisión Nuevo 2026-07-01.xlsx', data_only=True)
ws = wb.active
headers = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=6, column=c).value
    if h: headers[h.strip()] = c
c_dias = [c for h,c in headers.items() if 'Días Licencias' in h][0]
c_rut = [c for h,c in headers.items() if 'mero de Documento' in h][0]
c_nombre = [c for h,c in headers.items() if 'Nombre Completo' in h][0]

con_licencia = []
for r in range(7, ws.max_row+1):
    d = ws.cell(row=r, column=c_dias).value
    if d and d > 0:
        rut = ws.cell(row=r, column=c_rut).value
        nombre = ws.cell(row=r, column=c_nombre).value
        con_licencia.append({'rutNorm': norm_rut(str(rut)), 'nombre': nombre, 'dias': d})

ref = set(x['rutNorm'] for x in con_licencia)
print(f'trabajadores con licencia (LH): {len(ref)}')

txt_completo, total = get_pdf_text_ocr(r'Documentos Antofagasta/M).-Licencias Medicas/licencia medica.pdf', max_pages_ocr=100)
print(f'texto total extraido: {len(txt_completo)} chars')

# 1) filas tabulares
filas = VA_LIC_TABLA_ROW_RE.findall(txt_completo)
ruts_tabla = set(norm_rut(f[0]) for f in filas if norm_rut(f[0]) in ref)
print(f'filas tabulares encontradas: {len(filas)} | de la nomina de licencia: {len(ruts_tabla)}')

# 2) formulario individual (regex general)
ruts_generico = find_all_ruts(txt_completo)
coinc_generico = ruts_generico & ref

todos = ruts_tabla | coinc_generico
print(f'RUT encontrados (formulario general): {len(ruts_generico)} | coincidencia: {len(coinc_generico)}')
print(f'TOTAL cobertura (tabla + generico): {len(todos)}/{len(ref)} = {len(todos)/len(ref)*100:.1f}%')
print('faltantes:', ref - todos)
