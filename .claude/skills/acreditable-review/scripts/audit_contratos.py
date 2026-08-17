import sys, re
sys.path.insert(0, r'.claude\skills\acreditable-review\scripts')
from audit_antofagasta import *
import fitz
from PIL import Image
import io

VA_CONTRATO_TIPOS = ['CONTRATO','ANEXO_CARGO','ANEXO_HHEE']

def clasificar_titulo_contrato(txt):
    if re.search(r'ANEXO\s+DE\s+CONTRATO', txt, re.I):
        return 'ANEXO_HHEE' if re.search(r'Pacto\s+Horas\s+Extras|horas\s+extraordinarias', txt, re.I) else 'ANEXO_CARGO'
    if re.search(r'CONTRATO\s+DE\s+TRABAJO', txt, re.I):
        return 'CONTRATO'
    return None

def detectar_worker_en_titulo(txt, lista_workers):
    for r in find_all_ruts(txt):
        for w in lista_workers:
            if w['rutNorm'] == r:
                return w
    m = re.search(r'[Dd]on\s*\(?[ñn]?a?\)?\s*([A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúñ\s]+?),\s*de\s*nacionalidad', txt)
    if m:
        match, motivo, _ = match_nombre_nomina_con_motivo(m.group(1).strip(), lista_workers)
        if motivo == 'ok':
            return match
    return None

rows = cargar_lh(r'Documentos Antofagasta\2026-08-17 Libro de haberes Revisión Nuevo 2026-07-01.xlsx')
# nuevosLH: ingreso en julio (periodo del archivo) -- necesitamos "Fecha Ingreso Compañía"
import openpyxl
wb = openpyxl.load_workbook(r'Documentos Antofagasta\2026-08-17 Libro de haberes Revisión Nuevo 2026-07-01.xlsx', data_only=True)
ws = wb.active
headers = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=6, column=c).value
    if h: headers[h.strip()] = c
c_ingreso = [c for h,c in headers.items() if 'Fecha Ingreso Compa' in h][0]
c_rut = [c for h,c in headers.items() if 'mero de Documento' in h][0]
c_nombre = [c for h,c in headers.items() if 'Nombre Completo' in h][0]

nuevos = []
for r in range(7, ws.max_row+1):
    ing = ws.cell(row=r, column=c_ingreso).value
    if ing and ing.month == 7 and ing.year == 2026:
        rut = ws.cell(row=r, column=c_rut).value
        nombre = ws.cell(row=r, column=c_nombre).value
        nuevos.append({'rutNorm': norm_rut(str(rut)), 'rut': str(rut), 'nombre': nombre})

print(f'Trabajadores nuevos (ingreso julio 2026): {len(nuevos)}')
for n in nuevos:
    print(f"  {n['rut']} {n['nombre']}")

doc = fitz.open(r'Documentos Antofagasta/H).-Finiquitos - Seguro Cesantia - Contratos - Liquidaciones/Contratos de trabajo/Contratos de trabajos.pdf')
print(f'\npáginas: {doc.page_count}')

worker = None
docs = {}
def nuevos_docs():
    return {t: {'presente': False} for t in VA_CONTRATO_TIPOS}
docs = nuevos_docs()
resultados = {}  # rutNorm -> set(tipos presentes)
tipo_actual = None

for i in range(doc.page_count):
    page = doc[i]
    mat = fitz.Matrix(2.6, 2.6)
    pix = page.get_pixmap(matrix=mat)
    img = Image.open(io.BytesIO(pix.tobytes('png')))
    txt = pytesseract.image_to_string(img, lang='spa')
    tipo_pagina = clasificar_titulo_contrato(txt)
    if tipo_pagina:
        detectado = detectar_worker_en_titulo(txt, nuevos)
        if detectado and (not worker or detectado['rutNorm'] != worker['rutNorm']):
            worker = detectado
            resultados.setdefault(worker['rutNorm'], {'nombre': worker['nombre'], 'tipos': set()})
        tipo_actual = tipo_pagina
        print(f'pag {i+1}: TITULO={tipo_pagina} worker_detectado={"SI:"+detectado["nombre"] if detectado else "NO"} worker_actual={worker["nombre"] if worker else None}')
    if tipo_actual and worker:
        resultados[worker['rutNorm']]['tipos'].add(tipo_actual)

print('\n=== RESUMEN POR TRABAJADOR DETECTADO EN EL ARCHIVO ===')
for rutn, info in resultados.items():
    faltan = set(VA_CONTRATO_TIPOS) - info['tipos']
    print(f"  {info['nombre']}: tiene {sorted(info['tipos'])} | falta {sorted(faltan) if faltan else 'NADA -- completo'}")

detectados_ruts = set(resultados.keys())
nuevos_ruts = set(n['rutNorm'] for n in nuevos)
print(f'\nDe los {len(nuevos)} nuevos de julio, el archivo agrupado identificó a {len(detectados_ruts & nuevos_ruts)}')
print('nuevos NO detectados en el archivo (candidatos a falso "sin archivo"):')
for n in nuevos:
    if n['rutNorm'] not in detectados_ruts:
        print(f"  {n['rut']} {n['nombre']}")
