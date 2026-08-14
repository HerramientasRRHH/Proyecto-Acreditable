# -*- coding: utf-8 -*-
r"""
Réplica en Python de va_normalizarNombre / va_tokensNombre / va_coberturaTokens
/ va_matchNombreNomina (index.html, ~línea 10886+) — el matching de nombres
por tokens que reemplazó al Levenshtein-de-string-completo.

Usalo para probar en Python, ANTES de tocar el JS, si un nombre leído por la
IA matchearía o no contra la nómina real — mismo criterio, mismo resultado.
Si cambiás el algoritmo en index.html, actualizá esta copia también (o al
revés: si encontrás un caso que rompe acá, el fix probablemente va primero
acá y después se porta al JS).

Uso como módulo:
    from match_nombre import match_nueva, cargar_nomina
    nombres = cargar_nomina(r"C:\...\Libro de haberes....xlsx")
    resultado, motivo = match_nueva("Dasnia Arcos Salinas", nombres)

Uso como script (sanity check anti-falsos-positivos contra una nómina real):
    python match_nombre.py "ruta\Libro de haberes....xlsx"
"""
import sys
import unicodedata

import openpyxl

STOP = {"SR", "SRA", "DON", "DONA", "SENOR", "SENORA", "DE", "LA", "EL"}


def norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().upper()
    s = "".join(c if c.isalnum() or c == " " else " " for c in s)
    return " ".join(s.split())


def tokens(s):
    seen = []
    for t in norm(s).split(" "):
        if t and t not in STOP and len(t) >= 2 and t not in seen:
            seen.append(t)
    return seen


def levenshtein(a, b):
    if a == b:
        return 0
    la, lb = len(a), len(b)
    if la == 0:
        return lb
    if lb == 0:
        return la
    prev = list(range(lb + 1))
    for i in range(1, la + 1):
        cur = [i] + [0] * lb
        for j in range(1, lb + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            cur[j] = min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + cost)
        prev = cur
    return prev[lb]


def token_match(t1, t2):
    if t1 == t2:
        return True
    d = levenshtein(t1, t2)
    tol = 1 if min(len(t1), len(t2)) <= 5 else 2
    return d <= tol


def coverage(a_tokens, b_tokens):
    if not a_tokens:
        return 0.0
    matched = 0
    used = set()
    for lt in a_tokens:
        for i, ht in enumerate(b_tokens):
            if i in used:
                continue
            if token_match(lt, ht):
                matched += 1
                used.add(i)
                break
    return matched / len(a_tokens)


def match_nueva(nombre_leido, lista_lh):
    """Replica exacta de va_matchNombreNomina. lista_lh: lista de strings
    (nombres) o de objetos con atributo/clave 'nombre'."""
    def nombre_de(x):
        return x if isinstance(x, str) else x.get("nombre") if isinstance(x, dict) else x.nombre

    lt = tokens(nombre_leido)
    if len(lt) < 1:
        return None, "sin tokens utiles"
    scored = [(coverage(lt, tokens(nombre_de(t))), t) for t in lista_lh]
    scored.sort(key=lambda x: -x[0])
    mejor_cov, mejor = scored[0]
    segundo_cov = scored[1][0] if len(scored) > 1 else 0.0

    if len(lt) >= 2:
        if mejor_cov < 1.0:
            return None, f"rechazado: cobertura incompleta (mejor={nombre_de(mejor)!r} cov={mejor_cov:.2f})"
    else:
        if mejor_cov < 1.0 or len(lt[0]) < 4:
            return None, "rechazado: 1 solo token corto/incompleto"

    if mejor_cov - segundo_cov < 0.34:
        empatados = [t for c, t in scored if c >= mejor_cov - 1e-9]
        if len(empatados) > 1:
            return None, f"rechazado: ambiguo entre {[nombre_de(t) for t in empatados[:4]]}"

    return mejor, f"MATCH cov={mejor_cov:.2f} vs 2do={segundo_cov:.2f}"


def cargar_nomina(path_xlsx, incluir_desvinculados=False):
    """Lee el Libro de Haberes (headers en fila 6) y devuelve la lista de
    nombres completos activos (o todos, si incluir_desvinculados=True)."""
    wb = openpyxl.load_workbook(path_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [ws.cell(row=6, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(hdr) if h}
    nombres = []
    for r in range(7, ws.max_row + 1):
        nom = ws.cell(row=r, column=idx["Nombre Completo"]).value
        if not nom:
            continue
        if not incluir_desvinculados:
            termino = ws.cell(row=r, column=idx.get("Fecha Término Trabajo", 0)).value if "Fecha Término Trabajo" in idx else None
            if termino:
                continue
        nombres.append(nom.strip())
    return nombres


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Uso: python match_nombre.py \"ruta\\Libro de haberes....xlsx\"")
    nombres = cargar_nomina(sys.argv[1])
    print(f"Trabajadores activos: {len(nombres)}")
    print("Sanity check anti-falsos-positivos: cada nombre real probado contra si mismo...")
    errores = 0
    for real in nombres:
        r, why = match_nueva(real, nombres)
        if r != real:
            errores += 1
            print(f"  !! {real!r} -> matcheo con {r!r} en vez de si mismo  ({why})")
    print(f"Errores: {errores}/{len(nombres)}")
